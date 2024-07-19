import csv
import os
import io
import sys
import json
import math
import copy
import random
import logging
from re import U, template
import string
from datetime import datetime
import traceback
# from typing import OrderedDict
import uuid

from pprint import pprint
from flask.signals import before_render_template

import pandas as pd

from flask import request
from flask import flash
from flask import render_template
from flask import Blueprint
from flask import current_app
from flask import redirect
from flask import url_for
from flask import jsonify
from flask import send_file
from sqlalchemy.sql.expression import extract, select
from werkzeug import utils


from werkzeug.utils import secure_filename

from flask_login import login_required
from flask_login import current_user
from flask_login.utils import login_fresh

from lnma import dora
from lnma import util
from lnma import ss_state
from lnma import settings
from lnma import srv_paper
from lnma import srv_extract
from lnma.settings import TOXIXITY_ABBR_GROUP
from lnma import create_app, redis_client
import lnma

bp = Blueprint("extractor", __name__, url_prefix="/extractor")
template_base = 'extractor/'
app = create_app()
logger = app.logger

@bp.route('/v1')
@login_required
def v1():
    project_id = request.cookies.get('project_id')
    cq_abbr = request.cookies.get('cq_abbr')

    if project_id is None:
        return redirect(url_for('project.mylist'))

    if cq_abbr is None:
        cq_abbr = 'default'

    project = dora.get_project(project_id)

    return render_template(
        'extractor/v1.html',
        cq_abbr=cq_abbr,
        project=project,
        project_json_str=json.dumps(project.as_dict())
    )


@bp.route('/manage_outcomes')
@login_required
def manage_outcomes():
    project_id = request.cookies.get('project_id')
    cq_abbr = request.cookies.get('cq_abbr')

    if project_id is None:
        return redirect(url_for('project.mylist'))

    if cq_abbr is None:
        cq_abbr = 'default'

    project = dora.get_project(project_id)
    outcomes_updated = []
    cqs = project.settings['clinical_questions']
    for cq in cqs:
        if cq['abbr'] == cq_abbr:
            if cq.get('toxicity') and cq.get('toxicity') =='yes':
               outcomes_ = project.settings['outcome']
               outcomes_updated = []
               for abbr in cq['tox_groups']:
                   outcomes_updated.append({'abbr': abbr, 'name': TOXIXITY_ABBR_GROUP[abbr]})
               if (cq.get('parent_group') =='pwma'):
                   project.settings['outcome'] = {'pwma': {'analysis_groups': outcomes_updated}, 'nma': {'analysis_groups': []}}
               else:
                   project.settings['outcome'] = {'nma': {'analysis_groups': outcomes_updated}, 'pwma': {'analysis_groups': []}}
            
               project.settings['outcomes_enabled'] = [cq.get('parent_group')]

    return render_template(
        template_base + 'manage_outcomes.html',
        cq_abbr=cq_abbr,
        project=project,
        project_json_str=json.dumps(project.as_dict())
    )


@bp.route('/manage_itable')
@login_required
def manage_itable():
    '''
    Design the ITable
    '''
    project_id = request.cookies.get('project_id')

    if project_id is None:
        return redirect(url_for('project.mylist'))

    # decide which cq to use
    cq_abbr = request.cookies.get('cq_abbr')
    if cq_abbr is None:
        cq_abbr = 'default'

    project = dora.get_project(project_id)

    return render_template(
        template_base + 'manage_itable.html',
        cq_abbr=cq_abbr,
        project=project,
        project_json_str=json.dumps(project.as_dict())
    )


@bp.route('/result_table', methods=['GET'])
@login_required
def result_table():
    '''
    Design the Result Table 
    '''
    project_id = request.args.get('project_id')

    if project_id is None:
        return redirect(url_for('project.mylist'))

    # decide which cq to use
    cq_abbr = request.args.get('cq')
    if cq_abbr is None:
        cq_abbr = 'default'
    project = dora.get_project(project_id)

    keys_list = project.settings.get('clinical_question_result_table')
    list_ = []

    if project.settings.get('clinical_question_result_table'):
        if project.settings.get('clinical_question_result_table').get(cq_abbr):
            if project.settings.get('clinical_question_result_table').get(cq_abbr).get('cate_attrs_relations'):
                keys_list = project.settings.get('clinical_question_result_table').get(cq_abbr).get('cate_attrs_relations')
                for key in keys_list:
                    for i, j in key.items():
                        list_.append(i)
    

    if project.settings.get('clinical_question_result_table'):

        if project.settings.get('clinical_question_result_table').get(cq_abbr):
            return render_template(
            template_base + 'result_table.html',
            project_json_str=list_,
            cq_abbr=cq_abbr,
            project=project,
            comaparator_name=json.dumps(project.settings.get('clinical_question_result_table').get(cq_abbr).get('comaparator_name')),
            varaiable_name=json.dumps(project.settings.get('clinical_question_result_table').get(cq_abbr).get('varaiable_name')), 
            abbr_list_updated=json.dumps(project.settings.get('clinical_question_result_table').get(cq_abbr).get('abbr_list_updated')),
            headers_result=json.dumps(project.settings.get('clinical_question_result_table').get(cq_abbr).get('cate_attrs_relations')),
            rows_result=json.dumps(project.settings.get('clinical_question_result_table').get(cq_abbr).get('result_table_records')),
            rows_result_seq=json.dumps(project.settings.get('clinical_question_result_table').get(cq_abbr).get('result_table_records_seq')),
            rows_headers_mapping=json.dumps(project.settings.get('clinical_question_result_table').get(cq_abbr).get('parent_nested_realtion'))
            )
        else:
            return render_template(
            template_base + 'result_table.html',
            cq_abbr=cq_abbr,
            project=project,
            headers_result={},
            rows_result={},
            rows_headers_mapping={}
            )

    
    else:

        return render_template(
            template_base + 'result_table.html',
            cq_abbr=cq_abbr,
            project=project,
            headers_result={},
            rows_result={},
            rows_headers_mapping={}
        )


#     project = dora.get_project(project_id)
#     resut_table_json= {
#         "Categories": ["1 (IO)", "2(IO/IO)", "2(IO/TKI)", "3(IO/IO/TKI)"],
#         "Trial_relations" : {"1 (IO)": {"sub": ["Trial A", "Trial B"]}},
#         "1 (IO)": { "OS HR": "1" , "PFS HR ": "2" , "Median PFS , months": "3"},
#         "1 (IO)": { "OS HR": "1" , "PFS HR ": "2" , "Median PFS , months": "3"},
#         "1 (IO)": { "OS HR": "1" , "PFS HR ": "2" , "Median PFS , months": "3"}
#     }

#     result=  [
#   {"1 (IO)": [{"subs": [{"KEYNOTE-427(Pembrolizumab)²n=68": [{"sub_categories":{}}]} ]}] },
#   {"2(IO/IO)": [{"subs": [{"CheckMate 214(Ipilimumab/Nivolumab)³n=425": [{"sub_categories":{}}]} ]}] },
#   {"2(IO/TKI)": [{"subs": [{"KEYNOTE-426(Axitinib/Pembrolizumab4n=294": [{"sub_categories":{}}]},
#   {"CheckMate 9ER(Cabozantinib/Nivolumab)5 n=249": [{"sub_categories":{}}]},
#   {"CLEAR (Lenvatinib/Pembrolizumab)6 n=243": [{"sub_categories":{}}]}
  
#    ]}] },
#   {"3(IO/IO/TKI)": [{"subs": [{"COSMIC-313(Ipilimumab/Nivolumab/Cabozantinib)7n=428": [{"sub_categories":{}}]} ]}] }
  
#   ]

#     return render_template(
#         template_base + 'result_table.html',
#         cq_abbr=cq_abbr,
#         project=project,
#         project_json_str=resut_table_json,
#         result=json.dumps(result)
#     )


@bp.route('/manage_result_table', methods=['GET', 'POST'])
@login_required
def manage_result_table():
    '''
    Design the Result Table 
    '''
    if request.method == 'GET':
        project_id = request.cookies.get('project_id')

        if project_id is None:
            return redirect(url_for('project.mylist'))

        # decide which cq to use
        cq_abbr = request.cookies.get('cq_abbr')
        if cq_abbr is None:
            cq_abbr = 'default'
        project = dora.get_project(project_id)

        keys_list = project.settings.get('clinical_question_result_table')
        list_ = []

        if project.settings.get('clinical_question_result_table'):
            if project.settings.get('clinical_question_result_table').get(cq_abbr):
                if project.settings.get('clinical_question_result_table').get(cq_abbr).get('cate_attrs_relations'):
                    keys_list = project.settings.get('clinical_question_result_table').get(cq_abbr).get('cate_attrs_relations')
                    for key in keys_list:
                        for i, j in key.items():
                            list_.append(i)
        

        if project.settings.get('clinical_question_result_table'):

            if project.settings.get('clinical_question_result_table').get(cq_abbr):
                return render_template(
                template_base + 'manage_result_table.html',
                project_json_str=list_,
                cq_abbr=cq_abbr,
                project=project,
                comaparator_name=json.dumps(project.settings.get('clinical_question_result_table').get(cq_abbr).get('comaparator_name')),
                varaiable_name=json.dumps(project.settings.get('clinical_question_result_table').get(cq_abbr).get('varaiable_name')), 
                abbr_list_updated=json.dumps(project.settings.get('clinical_question_result_table').get(cq_abbr).get('abbr_list_updated')),
                headers_result=json.dumps(project.settings.get('clinical_question_result_table').get(cq_abbr).get('cate_attrs_relations')),
                rows_result=json.dumps(project.settings.get('clinical_question_result_table').get(cq_abbr).get('result_table_records')),
                rows_result_seq=json.dumps(project.settings.get('clinical_question_result_table').get(cq_abbr).get('result_table_records_seq')),
                rows_headers_mapping=json.dumps(project.settings.get('clinical_question_result_table').get(cq_abbr).get('parent_nested_realtion'))
                )
            else:
                return render_template(
                template_base + 'manage_result_table.html',
                cq_abbr=cq_abbr,
                project=project,
                headers_result={},
                rows_result={},
                rows_headers_mapping={}
                )

        
        else:

            return render_template(
                template_base + 'manage_result_table.html',
                cq_abbr=cq_abbr,
                project=project,
                headers_result={},
                rows_result={},
                rows_headers_mapping={}
            )


    else:
        project_id = request.form.get('project_id')
        cq_abbr = request.form.get('cq_abbr')
        meta = json.loads(request.form.get('result_table_headers'))
        
        project = dora.get_project(project_id)

        project_settings = project.settings 
        cq_result_table = project_settings.get('clinical_question_result_table')
        if cq_result_table.get(cq_abbr): 
            project_settings['clinical_question_result_table'][cq_abbr]['cate_attrs_relations'] = meta
            abbr_list = []
            seq_data = []
            for item in meta:
                abbr = item["abbr"]
                # abbr_list.append(abbr)
                dict_ ={}
                dict_name ={}
                for attr in item["attrs"]:
                    sub_abbr = attr["abbr"]
                    sub_name = attr['name']
                    dict_[sub_abbr] = []
                    dict_name[sub_abbr] = []
                    # abbr_list.append(f"{abbr} -> {sub_abbr}")
                    
                    for sub_attr in attr["sub_categories"]:
                        sub_sub_abbr = sub_attr["abbr"]
                        sub_sub_name = sub_attr['name']

                        dict_[sub_abbr].append(sub_sub_abbr)
                        
                        # dict_name[sub_abbr].append(sub_sub_name)
                        dict_name[sub_abbr].append(F'{sub_name} | {sub_sub_name}')

                        # abbr_list.append(f"{abbr} -> {sub_abbr} -> {sub_sub_abbr}")
                    if len(dict_[sub_abbr]):
                        for r in dict_[sub_abbr]:
                            abbr_list.append(f"{abbr} -> {sub_abbr} -> {r}")
             
                        for r in dict_name[sub_abbr]:
             
                            if sub_sub_name in seq_data:
                               seq_data.append(f'{r}.1')
                            else:
                                seq_data.append(r)
                    else:
                        abbr_list.append(f"{abbr} -> {sub_abbr}")
                        if sub_name in seq_data:
                            seq_data.append(f'{sub_name}.1')
                        else:
                            seq_data.append(sub_name)
            abbr_list_updated = [] 
            for item in abbr_list:
                parts = item.split(" -> ")
                abbr_list_updated.append(parts)
            project_settings['clinical_question_result_table'][cq_abbr]['abbr_list_updated'] = abbr_list_updated
            project_settings['clinical_question_result_table'][cq_abbr]['result_table_records_seq'] = seq_data
        

            is_success, project = dora.set_project_settings(
                project_id, project_settings
            )

        ret = {
            'success': True,
            'msg': 'Successfully Updated Result Table',
        }
        return jsonify(ret)



@bp.route('/manage_result_records', methods=['GET', 'POST'])
@login_required
def manage_result_records():
    '''
    Design the Result Table 
    '''
    if request.method == 'GET':
        project_id = request.cookies.get('project_id')

        if project_id is None:
            return redirect(url_for('project.mylist'))

        # decide which cq to use
        cq_abbr = request.cookies.get('cq_abbr')
        if cq_abbr is None:
            cq_abbr = 'default'
        project = dora.get_project(project_id)

        keys_list = project.settings.get('clinical_question_result_table')
        list_ = []

        if project.settings.get('clinical_question_result_table'):
            if project.settings.get('clinical_question_result_table').get(cq_abbr):
                if project.settings.get('clinical_question_result_table').get(cq_abbr).get('cate_attrs_relations'):
                    keys_list = project.settings.get('clinical_question_result_table').get(cq_abbr).get('cate_attrs_relations')
                    for key in keys_list:
                        for i, j in key.items():
                            list_.append(i)
        

        if project.settings.get('clinical_question_result_table'):

            if project.settings.get('clinical_question_result_table').get(cq_abbr):
                return render_template(
                template_base + 'manage_result_records.html',
                project_json_str=list_,
                cq_abbr=cq_abbr,
                project=project,
                comaparator_name=json.dumps(project.settings.get('clinical_question_result_table').get(cq_abbr).get('comaparator_name')),
                varaiable_name=json.dumps(project.settings.get('clinical_question_result_table').get(cq_abbr).get('varaiable_name')), 
                abbr_list_updated=json.dumps(project.settings.get('clinical_question_result_table').get(cq_abbr).get('abbr_list_updated')),
                headers_result=json.dumps(project.settings.get('clinical_question_result_table').get(cq_abbr).get('cate_attrs_relations')),
                rows_result=json.dumps(project.settings.get('clinical_question_result_table').get(cq_abbr).get('result_table_records')),
                rows_result_seq=json.dumps(project.settings.get('clinical_question_result_table').get(cq_abbr).get('result_table_records_seq')),
                rows_headers_mapping=json.dumps(project.settings.get('clinical_question_result_table').get(cq_abbr).get('parent_nested_realtion'))
                )
            else:
                return render_template(
                template_base + 'manage_result_records.html',
                cq_abbr=cq_abbr,
                project=project,
                headers_result={},
                rows_result={},
                rows_headers_mapping={}
                )

        
        else:

            return render_template(
                template_base + 'manage_result_records.html',
                cq_abbr=cq_abbr,
                project=project,
                headers_result={},
                rows_result={},
                rows_headers_mapping={}
            )


    else:
        project_id = request.form.get('project_id')
        cq_abbr = request.form.get('cq_abbr')
        meta = json.loads(request.form.get('result_table_rows'))
        
        project = dora.get_project(project_id)

        project_settings = project.settings 
        cq_result_table = project_settings.get('clinical_question_result_table')
        if cq_result_table.get(cq_abbr): 
            project_settings['clinical_question_result_table'][cq_abbr]['result_table_records'] = meta
        

            is_success, project = dora.set_project_settings(
                project_id, project_settings
            )

        ret = {
            'success': True,
            'msg': 'Successfully Updated Result Table',
        }
        return jsonify(ret)





@bp.route('/extract_by_paper')
@login_required
def extract_by_paper():
    project_id = request.cookies.get('project_id')

    if project_id is None:
        return redirect(url_for('project.mylist'))

    # decide which cq to use
    cq_abbr = request.cookies.get('cq_abbr')
    if cq_abbr is None:
        cq_abbr = 'default'

    project = dora.get_project(project_id)

    return render_template(
        template_base + 'extract_by_paper.html',
        cq_abbr=cq_abbr,
        project=project,
        project_json_str=json.dumps(project.as_dict())
    )


@bp.route('/extract_by_outcome')
@login_required
def extract_by_outcome():
    '''
    Extract data
    '''
    project_id = request.cookies.get('project_id')
    # project_id = request.args.get('project_id')
    if project_id is None:
        return redirect(url_for('project.mylist'))

    # decide which cq to use
    cq_abbr = request.cookies.get('cq_abbr')
    if cq_abbr is None:
        cq_abbr = 'default'

    oc_abbr = request.args.get('abbr')
    project = dora.get_project(project_id)

    return render_template(
        template_base + 'extract_by_outcome.html', 
        oc_abbr=oc_abbr,
        cq_abbr=cq_abbr,
        project=project,
        project_json_str=json.dumps(project.as_dict())
    )


@bp.route('/check_outcome_quality')
@login_required
def check_outcome_quality():
    '''
    Check outcome quality
    '''
    project_id = request.cookies.get('project_id')
    # project_id = request.args.get('project_id')
    if project_id is None:
        return redirect(url_for('project.mylist'))

    # decide which cq to use
    cq_abbr = request.cookies.get('cq_abbr')
    if cq_abbr is None:
        cq_abbr = 'default'

    oc_abbr = request.args.get('abbr')
    project = dora.get_project(project_id)

    return render_template(
        template_base + 'check_outcome_quality.html', 
        oc_abbr=oc_abbr,
        cq_abbr=cq_abbr,
        project=project,
        project_json_str=json.dumps(project.as_dict())
    )


@bp.route('/check_data_quality')
@login_required
def check_data_quality():
    '''
    Check data quality
    '''
    project_id = request.cookies.get('project_id')
    # project_id = request.args.get('project_id')
    if project_id is None:
        return redirect(url_for('project.mylist'))

    # decide which cq to use
    cq_abbr = request.cookies.get('cq_abbr')
    if cq_abbr is None:
        cq_abbr = 'default'

    oc_abbr = request.args.get('abbr')
    project = dora.get_project(project_id)

    return render_template(
        template_base + 'check_data_quality.html', 
        oc_abbr=oc_abbr,
        cq_abbr=cq_abbr,
        project=project,
        project_json_str=json.dumps(project.as_dict())
    )


@bp.route('/get_duplicate_outcomes')
@login_required
def get_duplicate_outcomes():
    '''
    Get duplicate outcomes
    '''
    project_id = request.args.get('project_id')

    if project_id is None:
        ret = {
            'success': False,
            'msg': 'project_id is required'
        }
        return jsonify(ret)
    project = dora.get_project(project_id)

    # decide which cq to use
    cq_abbr = request.args.get('cq_abbr')

    # get all extracts of this project/cq
    extracts = dora.get_extracts_by_project_id_and_cq(
        project_id,
        cq_abbr
    )

    # try to get concept_synonyms
    # if not defined, just get {}
    concept_synonyms = project.settings.get(
        'concept_synonyms',
        {}
    )
    concept_dict = util.get_concept_dict_from_concept_synonyms(
        concept_synonyms
    )

    # get all duplicates
    dup_ocs = srv_extract.get_duplicate_outcomes_from_extracts(
        extracts,
        concept_dict
    )

    dup_ocs_stat = srv_extract.get_stat_outcomes(
        extracts
    )

    # build output
    json_extracts = [ extr.as_simple_dict() for extr in extracts ]

    ret = {
        'success': True,
        'data': {
            'report': dup_ocs,
            'stat': dup_ocs_stat,
            'extracts': json_extracts,
        },
        'last_checked': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }

    return jsonify(ret)


@bp.route('/get_data_quality_report')
@login_required
def get_data_quality_report():
    '''
    Check data quality
    '''
    project_id = request.args.get('project_id')

    if project_id is None:
        ret = {
            'success': False,
            'msg': 'project_id is required'
        }
        return jsonify(ret)
    project = dora.get_project(project_id)

    # decide which cq to use
    cq_abbr = request.args.get('cq_abbr')

    src = request.args.get('src')
    if src is None or src == '':
        # set the default to get things from db
        src = 'db'

    output_fn = 'DATA_QUALITY.json'
    full_output_fd = os.path.join(
        current_app.instance_path, 
        settings.PUBLIC_PATH_PUBDATA, 
        project.keystr, 
        cq_abbr
    )
    if not os.path.exists(full_output_fd):
        os.makedirs(full_output_fd, exist_ok=True)
        
    full_output_fn = os.path.join(
        full_output_fd,
        output_fn
    )
    if src == 'cache':
        ret = json.load(open(full_output_fn))
        return jsonify(ret)

    # get report 
    report = srv_extract.get_data_quality(project_id, cq_abbr)

    # get papers and extracts?
    papers = srv_paper.get_included_papers_by_cq(
        project_id, 
        cq_abbr
    )

    extracts = dora.get_extracts_by_project_id_and_cq(
        project_id,
        cq_abbr
    )

    json_papers = [ p.as_very_simple_dict() for p in papers ]
    json_extracts = [ extr.as_simple_dict() for extr in extracts ]

    ret = {
        'success': True,
        'last_checked': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'data': {
            'report': report,
            'papers': json_papers,
            'extracts': json_extracts
        }
    }
    # cache
    json.dump(ret, open(full_output_fn, 'w'), default=util.json_encoder)

    return jsonify(ret)


@bp.route('/extract_coe')
@login_required
def extract_coe():
    '''
    Extract CoE
    '''
    project_id = request.cookies.get('project_id')
    # project_id = request.args.get('project_id')
    if project_id is None:
        return redirect(url_for('project.mylist'))

    # decide which cq to use
    cq_abbr = request.cookies.get('cq_abbr')
    if cq_abbr is None:
        cq_abbr = 'default'

    oc_abbr = request.args.get('abbr')

    project = dora.get_project(project_id)

    return render_template(
        template_base + 'extract_coe.html', 
        oc_abbr=oc_abbr,
        cq_abbr=cq_abbr,
        project=project,
        project_json_str=json.dumps(project.as_dict())
    )


@bp.route('/upload_xls_files')
@login_required
def upload_xls_files():
    '''
    Uploaded XLS files for itable/pwma/nma
    '''
    project_id = request.cookies.get('project_id')
    # project_id = request.args.get('project_id')
    if project_id is None:
        return redirect(url_for('project.mylist'))

    # decide which cq to use
    cq_abbr = request.cookies.get('cq_abbr')
    if cq_abbr is None:
        cq_abbr = 'default'

    project = dora.get_project(project_id)

    return render_template(
        template_base + 'upload_xls_files.html', 
        cq_abbr=cq_abbr,
        project=project,
        project_json_str=json.dumps(project.as_dict())
    )


###########################################################
# APIs for extraction
###########################################################

@bp.route('/get_fill_wp_attrs')
@login_required
def get_fill_wp_attrs():
    '''
    Get working paper auto fill

    The auto fill is a tool copy the text from itable to an outcome.
    So, in the backend, we need to get the followings:

    1. the `from` and `to` pairs for a specific project
    2. the `pids` which are working on
    2. search the itable extraction and find the text
    
    But this is very difficult.
    '''

    return ''



@bp.route('/get_paper')
@login_required
def get_paper():
    project_id = request.args.get('project_id')
    project_id = request.cookies.get('project_id')
    pid = request.args.get('pid')

    paper = dora.get_paper_by_project_id_and_pid(project_id, pid)

    if paper is None:
        ret = {
            'success': False,
            'paper': None
        }
    else:
        json_paper = paper.as_dict()

        ret = {
            'success': True,
            'paper': json_paper
        }

    return jsonify(ret)


@bp.route('/get_papers_by_stage')
@login_required
def get_papers_by_stage():
    project_id = request.args.get('project_id')
    stage = request.args.get('stage')
    papers = dora.get_papers_by_stage(project_id, stage)
    json_papers = [ p.as_very_simple_dict() for p in papers ]

    ret = {
        'success': True,
        'msg': '',
        'papers': json_papers
    }
    return jsonify(ret)



@bp.route('/download_extract_rs_csv')
@login_required
def download_extract_rs_csv():
    '''
    Get the extracted records 
    '''
    extract_id = request.args.get('extract_id')

    if extract_id is None:
        return 'extract_id is required'

    # get the extract by its id
    extract = dora.get_extract(extract_id)

    if extract is None:
        return 'no such extract'

    papers = srv_paper.get_included_papers_by_cq(
        extract.project_id, 
        extract.meta['cq_abbr']
    )

    # make a dictionary for lookup
    paper_dict = {}
    for paper in papers:
        paper_dict[paper.pid] = paper

    # get raw rs
    rscfg = extract.get_raw_rs_cfg(
        paper_dict,
        True
    )

    # just need the rs
    rs = rscfg['rs']

    if len(rs) == 0:
        return ''

    # get the columns
    columns = rs[0].keys()

    print("* download_extract_rs_csv rscfg: %s" % (rs))

    # build df
    import io
    import csv
    output = io.StringIO()

    writer = csv.DictWriter(
        output, 
        fieldnames=columns, 
        delimiter=','
    )
    writer.writeheader()
    writer.writerows(rs)
    txt = output.getvalue()

    return txt


@bp.route('/update_paper_one_selection', methods=['POST'])
@login_required
def update_paper_one_selection():
    '''
    Update just one selection for a paper
    '''
    project_id = request.form.get('project_id')
    cq_abbr = request.form.get('cq_abbr')
    pid = request.form.get('pid')
    abbr = request.form.get('abbr')
    is_selected = request.form.get('is_selected')
    is_selected = True if is_selected.lower() == 'true' else False

    # update status
    paper, extract, piece = dora.update_paper_selection(
        project_id, 
        cq_abbr,
        pid, 
        abbr, 
        is_selected
    )

    msg = 'Updated %s %s %s' % (
        paper.get_short_name(),
        'selected in' if is_selected else 'removed from',
        extract.get_short_title()
    )

    ret = {
        'success': True,
        'msg': msg,
        'data': {
            'cq_abbr': cq_abbr,
            'pid': pid,
            'abbr': abbr,
            'is_selected': is_selected,
            'piece': piece.as_dict()
        }
    }

    return jsonify(ret)


@bp.route('/update_paper_selections', methods=['POST'])
@login_required
def update_paper_selections():
    project_id = request.form.get('project_id')
    cq_abbr = request.form.get('cq_abbr')
    abbrs = request.form.getlist('abbrs[]')
    pid = request.form.get('pid')

    # update the extract with given info
    paper, outcome_selections = dora.update_paper_selections(
        project_id, cq_abbr, pid, abbrs
    )

    # bind the outcome selections to meta
    paper.meta['outcome_selections'] = outcome_selections

    # build the return obj
    ret = {
        'success': True,
        'msg': '',
        'paper': paper.as_dict()
    }
    return jsonify(ret)


@bp.route('/create_extract', methods=['POST'])
@login_required
def create_extract():
    project_id = request.form.get('project_id')
    oc_type = request.form.get('oc_type')
    abbr = request.form.get('abbr')
    meta = json.loads(request.form.get('meta'))
    data = json.loads(request.form.get('data'))

    # check if exists first
    extract = dora.get_extract_by_project_id_and_abbr(
        project_id, abbr
    )
    if extract is not None:
        # if get the exisiting extracts
        ret = {
            'success': False,
            'msg': 'exists extract %s' % abbr,
            'extract': extract.as_dict()
        }
        return jsonify(ret)

    # ok, not exists! create!
    extract = dora.create_extract(
        project_id,
        oc_type,
        abbr,
        meta,
        data
    )

    # build the return obj
    ret = {
        'success': True,
        'msg': '',
        'extract': extract.as_dict()
    }
    return jsonify(ret)



@bp.route('/get_extract_and_papers')
@login_required
def get_extract_and_papers():
    project_id = request.args.get('project_id')
    cq_abbr = request.args.get('cq_abbr')
    abbr = request.args.get('abbr')
    
    # get the exisiting extracts
    extract = dora.get_extract_by_project_id_and_abbr(
        project_id, 
        abbr
    )

    if extract is None:
        # this is a new extract
        ret = {
            'success': False,
            'msg': 'not exist extract %s' % abbr
        }
        return jsonify(ret)

    # 2023-04-06: fix None extract
    old_data = copy.deepcopy(extract.data)

    # get papers
    papers = srv_paper.get_included_papers_by_cq(
        project_id, 
        cq_abbr
    )
    paper_pids = [ p.pid for p in papers ]
    
    # update the meta
    extract.update_meta()

    # 2023-01-28: use piece to fill the data
    extract = dora.attach_extract_data(extract)
    
    extract_itable = srv_extract.get_itable_by_project_id_and_cq_abbr(
        project_id, cq_abbr
    )
    # if extract is None:
    #     return None
    
    # 2023-03-30: add data
    
    updated_trial_name_dict = {}
    abbr2attr_name = {} 
    paper_dict = {}
    if extract_itable:
        extract_itable = dora.attach_extract_data(extract_itable)
        papers_itable = dora.get_papers_of_included_sr(project_id)
        for p in papers_itable:
            paper_dict[p.pid] = p
            
        meta = extract_itable.meta
        data = extract_itable.data


        for cate in meta['cate_attrs']:
            for attr in cate['attrs']:
                if attr['subs'] is None or attr['subs'] == []:
                    attr_name = "%s" % attr['name']

                    abbr2attr_name[attr['abbr']] = attr_name
                    
                    # if(attr['name'].split() == 'Trial Name'.split()):
                    #     trial_name_abbr = attr['abbr']
                    # else:
                    #     attrs.append({
                    #         "attr_id": attr_id, 
                    #         "branch": "%s" % attr['name'], 
                    #         "cate": "%s" % cate['name'], 
                    #         "name": attr_name, 
                    #         "trunk": "_%s" % attr['name'], 
                    #         "vtype": "text",
                    #     })
        basic_attrs = [
        { 'abbr': 'ba_nct', 'name': 'NCT', },
    ]

        for attr in basic_attrs:
            attr_name = "%s" % attr['name']
            abbr2attr_name[attr['abbr']] = attr_name
        
        rs = []
        for pid in data:
            paper_ext = data[pid]

            if paper_ext['is_selected'] == False: continue

            # now, let's parse this paper
            # paper = dora.get_paper_by_project_id_and_pid(project_id, pid)
            if pid not in paper_dict:
                # what???
                print('* MISSING %s when building ITABLE.json' % pid)
                continue

            # get this paper from dict instead of SQL
            paper = paper_dict[pid]

            
            # the `r` is for the output
            # first, let's get the main arm
            r = dora._make_r_nct_based(
                paper_ext['attrs']['main']['g0'], 
                abbr2attr_name, 
                basic_attrs,
                paper
        
            )
            rs.append(r)
    
        for r in rs:
            if r['NCT']:
                updated_trial_name_dict[r['NCT']] = r.get('Trial Name', '') 

    # 2023-04-06: fix the removed data issue
    # It's possible that some studies are excluded after extraction
    # So, need to double check the pids
    missing_pids = []
    for pid in extract.data:
        if pid in paper_pids:
            # ok
            pass
        else:
            # oh ...
            missing_pids.append(pid)
    # remove those pids
    for pid in missing_pids:
        del extract.data[pid]

    # update the extract with papers
    # extract.update_data_by_papers(papers)
    extract_json = extract.as_dict()
    extract_json['old_data'] = old_data

    # make the return object
    from lnma.settings import STANDARD_DATA_COLS
    ret = {
        'success': True,
        'msg': '',
        'extract': extract_json,
        'papers': [ p.as_very_simple_dict() for p in papers ],
        'standard_data_cols':  STANDARD_DATA_COLS,
        'trial_name_dict': updated_trial_name_dict
    }
    return jsonify(ret)


@bp.route('/update_extract', methods=['POST'])
@login_required
def update_extract():
    '''
    Deprecated
    Update the extract meta and data
    '''
    project_id = request.form.get('project_id')
    oc_type = request.form.get('oc_type')
    abbr = request.form.get('abbr')

    # the meta of the extract settings
    meta = json.loads(request.form.get('meta'))

    # the data of the extracted infos
    data = json.loads(request.form.get('data'))
    
    # update the extract with given info
    extract = dora.update_extract_meta_and_data(
        project_id, oc_type, abbr, meta, data
    )

    # build the return obj
    ret = {
        'success': True,
        'msg': '',
        'extract': extract.as_dict()
    }
    return jsonify(ret)


@bp.route('/update_extract_meta', methods=['POST'])
@login_required
def update_extract_meta():
    '''
    Update the extract meta
    '''
    project_id = request.form.get('project_id')
    oc_type = request.form.get('oc_type')
    abbr = request.form.get('abbr')
    # the meta of the extract settings
    meta = json.loads(request.form.get('meta'))
    
    # update the extract with given info
    extract = dora.update_extract_meta(
        project_id, oc_type, abbr, meta
    )

    # build the return obj
    ret = {
        'success': True,
        'msg': '',
        'extract': extract.as_dict()
    }
    return jsonify(ret)


@bp.route('/update_extract_data', methods=['POST'])
@login_required
def update_extract_data():
    '''
    Update the extract data
    '''
    project_id = request.form.get('project_id')
    oc_type = request.form.get('oc_type')
    abbr = request.form.get('abbr')
    # the meta of the extract settings
    data = json.loads(request.form.get('data'))
    
    # update the extract with given info
    extract = dora.update_extract_data(
        project_id, oc_type, abbr, data
    )

    # build the return obj
    ret = {
        'success': True,
        'msg': '',
        'extract': extract.as_dict()
    }
    return jsonify(ret)


@bp.route('/update_extract_incr_data', methods=['POST'])
@login_required
def update_extract_incr_data():
    '''
    Update the extract data
    '''
    project_id = request.form.get('project_id')
    oc_type = request.form.get('oc_type')
    extract_id = request.form.get('extract_id')
    flag_skip_is_selected = request.form.get('flag_sis') == 'yes'
    
    # the meta of the extract settings
    data = json.loads(request.form.get('data'))
    
    # update the extract with given info
    extract, pieces = dora.update_extract_incr_data(
        extract_id, 
        data,
        flag_skip_is_selected
    )

    # build the return obj
    ret = {
        'success': True,
        'msg': '',
        'extract': extract.as_simple_dict()
    }
    return jsonify(ret)


@bp.route('/update_extract_coe_meta', methods=['POST'])
@login_required
def update_extract_coe_meta():
    '''
    Update the extract coe data
    '''
    extract_id = request.form.get('extract_id')
    
    # the meta of the extract settings
    coe = json.loads(request.form.get('coe'))
    
    # update the extract with given info
    extract = dora.update_extract_coe_meta(
        extract_id,
        coe
    )

    # build the return obj
    ret = {
        'success': True,
        'msg': '',
        'extract': extract.as_dict()
    }
    return jsonify(ret)


@bp.route('/copy_extract', methods=['POST'])
@login_required
def copy_extract():
    project_id = request.form.get('project_id')
    oc_type = request.form.get('oc_type')
    abbr = request.form.get('abbr')
    # the meta of the extract settings
    meta = json.loads(request.form.get('meta'))
    # the data of the extracted infos
    data = json.loads(request.form.get('data'))

    # the new extract
    new_abbr = request.form.get('new_abbr')
    new_full_name = request.form.get('new_full_name')
    
    # update the extract with given info
    _ = dora.update_extract_meta_and_data(
        project_id, oc_type, abbr, meta, data
    )

    # update the meta
    meta['abbr'] = new_abbr
    meta['full_name'] = new_full_name

    # save the extract with given info
    extract = dora.create_extract(
        project_id, oc_type, new_abbr, meta, data
    )

    # build the return obj
    ret = {
        'success': True,
        'msg': '',
        'extract': extract.as_dict()
    }
    return jsonify(ret)


@bp.route('/delete_extract', methods=['POST'])
@login_required
def delete_extract():
    project_id = request.form.get('project_id')
    oc_type = request.form.get('oc_type')
    abbr = request.form.get('abbr')
    
    # get the exisiting extracts
    _ = dora.delete_extract(project_id, oc_type, abbr)

    # build the return obj
    ret = {
        'success': True,
        'msg': ''
    }
    return jsonify(ret)


@bp.route('/get_pdata_in_extract')
@login_required
def get_pdata_in_extract():
    '''
    Get one paper in an extract by the project_id and the abbr and the pid
    '''
    project_id = request.args.get('project_id')
    abbr = request.args.get('abbr')
    pid = request.args.get('pid')
    
    # get the exisiting extracts
    extract = dora.get_extract_by_project_id_and_abbr(project_id, abbr)

    if extract is None:
        # this is a new extract
        ret = {
            'success': False,
            'abbr': abbr,
            'pid': pid,
            'msg': 'not exist extract %s' % abbr
        }
        return jsonify(ret)

    if pid not in extract.data:
        # this is a missing?
        ret = {
            'success': False,
            'abbr': abbr,
            'pid': pid,
            'msg': 'no paper data in itable'
        }
        return jsonify(ret)

    ret = {
        'success': True,
        'msg': '',
        'extract': extract.as_dict(),
    }

    # just keep the pid in extract
    ret['extract']['data'] = {
        pid: ret['extract']['data'][pid]
    }

    return jsonify(ret)


@bp.route('/get_pdata_in_itable')
@login_required
def get_pdata_in_itable():
    '''
    Get one paper in an itable by the project_id and the cq_abbr and the pid
    '''
    project_id = request.args.get('project_id')
    cq_abbr = request.args.get('cq_abbr')
    pid = request.args.get('pid')
    
    # get the exisiting extracts
    extract = srv_extract.get_itable_by_project_id_and_cq_abbr(
        project_id, cq_abbr
    )

    if extract is None:
        # this is a new extract
        ret = {
            'success': False,
            'abbr': None,
            'cq_abbr': cq_abbr,
            'pid': pid,
            'msg': 'not exist itable in cq %s' % cq_abbr
        }
        return jsonify(ret)

    if pid not in extract.data:
        # this is a missing?
        ret = {
            'success': False,
            'abbr': extract.abbr,
            'cq_abbr': cq_abbr,
            'pid': pid,
            'msg': 'no paper data in itable in cq %s' % cq_abbr
        }
        return jsonify(ret)

    ret = {
        'success': True,
        'msg': '',
        'extract': extract.as_dict(),
    }

    # just keep only ONE pid in extract
    # we don't need to send all results to user
    ret['extract']['data'] = {
        pid: ret['extract']['data'][pid]
    }

    return jsonify(ret)


@bp.route('/get_itable')
@login_required
def get_itable():
    '''
    Get one extract by the project_id and the abbr
    '''
    project_id = request.args.get('project_id')
    cq_abbr = request.args.get('cq_abbr')
    
    # get the exisiting extracts
    extract = dora.get_itable_by_project_id_and_cq(
        project_id, 
        cq_abbr
    )

    if extract is None:
        # this is a new extract
        ret = {
            'success': False,
            'msg': 'not exist itable for cq %s' % cq_abbr
        }
        return jsonify(ret)

    ret = {
        'success': True,
        'msg': '',
        'extract': extract.as_dict()
    }
    return jsonify(ret)


@bp.route('/get_extract_by_id')
@login_required
def get_extract_by_id():
    '''
    Get one extract by the extract_id
    '''
    extract_id = request.args.get('extract_id')
    
    # get the exisiting extracts
    extract = dora.get_extract(
        extract_id
    )

    if extract is None:
        # this is a new extract
        ret = {
            'success': False,
            'msg': 'not exist extract %s' % extract_id
        }
        return jsonify(ret)

    ret = {
        'success': True,
        'msg': '',
        'extract': extract.as_dict()
    }
    return jsonify(ret)


@bp.route('/get_extract')
@login_required
def get_extract():
    '''
    Get one extract by the project_id and the abbr
    '''
    project_id = request.args.get('project_id')
    cq_abbr = request.args.get('cq_abbr')
    abbr = request.args.get('abbr')
    
    # get the exisiting extracts
    extract = dora.get_extract_by_project_id_and_cq_and_abbr(
        project_id, 
        cq_abbr, 
        abbr
    )

    if extract is None:
        # this is a new extract
        ret = {
            'success': False,
            'msg': 'not exist extract %s' % abbr
        }
        return jsonify(ret)

    ret = {
        'success': True,
        'msg': '',
        'extract': extract.as_dict()
    }
    return jsonify(ret)


@bp.route('/get_extracts')
@login_required
def get_extracts():
    '''
    Get all of the extracts by the project_id
    '''
    project_id = request.args.get('project_id')

    if project_id is None:
        ret = {
            'success': False,
            'msg': 'project_id is required'
        }
        return jsonify(ret)
    project = dora.get_project(project_id)

    # with_data = request.args.get('with_data')
    # if with_data == 'yes':
    #     with_data = True
    # else:
    #     with_data = False
    
    # decide which cq to use
    # cq_abbr = request.cookies.get('cq_abbr')
    cq_abbr = request.args.get('cq_abbr')

    if cq_abbr is None:
        cq_abbr = 'default'
    
    # get the exisiting extracts
    extracts = dora.get_extracts_by_project_id_and_cq(
        project_id,
        cq_abbr
    )
    
    # update all data
    # extracts = dora.attach_all_extracts_data(extracts)
    papers = srv_paper.get_included_papers_by_cq(
        project_id, 
        cq_abbr
    )

    # update the selection relations
    papers, extracts = dora.update_papers_outcome_selections(
        project,
        papers,
        extracts
    )
    
    for i in range(len(extracts)):
        extracts[i].update_meta()

    # build the return obj

    ret = {
        'success': True,
        'msg': '',
        'extracts': [ extr.as_simple_dict() for extr in extracts ]
    }
        
    return jsonify(ret)



@bp.route('/get_extracts_bulk')

def get_extracts_bulk():
    '''
    Get all of the extracts by the project_id
    '''
    project_id = request.args.get('project_id')

    if project_id is None:
        ret = {
            'success': False,
            'msg': 'project_id is required'
        }
        return jsonify(ret)
    project = dora.get_project(project_id)

    # with_data = request.args.get('with_data')
    # if with_data == 'yes':
    #     with_data = True
    # else:
    #     with_data = False
    
    # decide which cq to use
    # cq_abbr = request.cookies.get('cq_abbr')
    cq_abbr = request.args.get('cq_abbr')

    if cq_abbr is None:
        cq_abbr = 'default'
    
    # get the exisiting extracts
    extracts = dora.get_extracts_by_project_id_and_cq(
        project_id,
        cq_abbr
    )
    
    # update all data
    # extracts = dora.attach_all_extracts_data(extracts)
    papers = srv_paper.get_included_papers_by_cq(
        project_id, 
        cq_abbr
    )
    papers_dict_seq = {}
    papers_dict_nct = {}
    for p_ in papers:
        papers_dict_seq[p_.pid] = p_.seq_num
        papers_dict_nct[p_.pid] = p_.get_rct_id()
        
    # update the selection relations
    papers, extracts = dora.update_papers_outcome_selections(
        project,
        papers,
        extracts
    )
    extracts_data = {}
    for i in range(len(extracts)):
       
        data_ = dora.get_attached_extract_data(extracts[i])
      
        # extracts[i]['new_attr'] ="abc" 
        extracts[i].update_meta()
        extracts_data[extracts[i].meta['full_name']] = data_
       
    n_selcted = len(extracts_data.keys())
    csv_header = ['', '', '', '', '', '', 'Regimen', '', 'All Grade', '', '', '', 'Grade3/4', '', 'Grade 3 or higher', '', 'Grade 5 only', '']
    csv_second_header = ['Seq_number', 'Trial Name', 'NCT', 'PMID', 'f(N selected)', 'Arms', 'Treatment', 'Control', 'Et', 'Nt', 'Ec', 'Nc', 'Et', 'Ec', 'Et', 'Ec', 'Et', 'Ec' ]
    sheet_data = []
    for key, value in extracts_data.items():
        sheet_dict = {}
        sheet_dict['sheet_name'] = key
        sheet_dict['headers'] = csv_header
        extracted_data = [] 
        extracted_data.append(csv_second_header)
        for key_, value_ in value.items():
            # Extract data from "attrs" -> "main" -> "g0"
            seq_num = papers_dict_seq.get(key_, '')
            nct_num = papers_dict_nct.get(key_, '')
            if "attrs" in value_ and "main" in value_["attrs"] and "g0" in value_["attrs"]["main"]:
                data__ = value_["attrs"]["main"]["g0"]
                extracted_data.append( [seq_num, '',nct_num, key_, '', value_.get('n_arms', ''),  
                data__.get('treatment', ''), data__.get('control', ''),
                data__.get('GA_Et', ''), data__.get('GA_Nt', ''),
                data__.get('GA_Ec', ''), data__.get('GA_Nc', ''),
                data__.get('G34_Et', ''), data__.get('G34_Et', ''),
                data__.get('G3H_Et', ''), data__.get('G3H_Ec', ''), 
               data__.get('G5N_Et', ''), data__.get('G5N_Ec', '')] )

            if value_['attrs']['other']:
                data__ = value_["attrs"]["other"][0]['g0']
                extracted_data.append( ['', '', '', '', '', '', 
                data__.get('treatment', ''), data__.get('control', ''),
                data__.get('GA_Et', ''), data__.get('GA_Nt', ''),
                data__.get('GA_Ec', ''), data__.get('GA_Nc', ''),
                data__.get('G34_Et', ''), data__.get('G34_Et', ''),
                data__.get('G3H_Et', ''), data__.get('G3H_Ec', ''), 
                data__.get('G5N_Et', ''), data__.get('G5N_Ec', '')] )
                if len(value_["attrs"]["other"])>1:
                    data__ = value_["attrs"]["other"][1]['g0']
                    extracted_data.append( ['', '', '', '', '', '', 
                    data__.get('treatment', ''), data__.get('control', ''),
                    data__.get('GA_Et', ''), data__.get('GA_Nt', ''),
                    data__.get('GA_Ec', ''), data__.get('GA_Nc', ''),
                    data__.get('G34_Et', ''), data__.get('G34_Et', ''),
                    data__.get('G3H_Et', ''), data__.get('G3H_Ec', ''), 
                data__.get('G5N_Et', ''), data__.get('G5N_Ec', '')] )





        sheet_dict['data'] = extracted_data
        sheet_data.append(sheet_dict)

    
    # return response
    from openpyxl import Workbook
    wb = Workbook()

    # Remove default sheet created with workbook
    wb.remove(wb.active)

    

    # Create each sheet and populate with data
    for sheet_info in sheet_data:
        sheet_name = sheet_info['sheet_name']
        headers = sheet_info['headers']
        data = sheet_info['data']

        # Create a new sheet
        sheet = wb.create_sheet(title=sheet_name)

        # Add headers
        sheet.append(headers)

       
        for row in data:
            print(row )
            sheet.append(row)

    # Save the workbook to a temporary file
    file_path = '/tmp/output.xlsx'  # Replace with your desired path
    wb.save(file_path)

    # Serve the file for download
    return send_file(file_path, as_attachment=True)


@bp.route('/get_extracts_io')

def get_extracts_io():
    '''
    Get all of the extracts by the project_id
    '''
    project_id = request.args.get('project_id')

    if project_id is None:
        ret = {
            'success': False,
            'msg': 'project_id is required'
        }
        return jsonify(ret)
    project = dora.get_project(project_id)

    cq_abbr = request.args.get('cq_abbr')
    abbr = request.args.get('abbr')
    select_extract = {}

    if cq_abbr is None:
        cq_abbr = 'default'
    
    # get the exisiting extracts
    extracts = dora.get_extracts_by_project_id_and_cq(
        project_id,
        cq_abbr
    )
    
    # update all data
    # extracts = dora.attach_all_extracts_data(extracts)
    papers = srv_paper.get_included_papers_by_cq(
        project_id, 
        cq_abbr
    )

    # update the selection relations
    papers, extracts = dora.update_papers_outcome_selections(
        project,
        papers,
        extracts
    )
    extracts_data = {}
    for i in range(len(extracts)):
       
        data_ = dora.get_attached_extract_data(extracts[i])
      
        # extracts[i]['new_attr'] ="abc" 
        extracts[i].update_meta()
        extracts_data[extracts[i].meta['full_name']] = data_
        if extracts[i].abbr == abbr:
            select_extract = extract[i]
            break
    n_selcted = len(extracts_data.keys())
    csv_header = ['', '', '', '', '', 'Regimen', '', 'All Grade', '', '', '', 'Grade3/4', '', 'Grade 3 or higher', '', 'Grade 5 only', '']
    csv_second_header = ['Trial Name', 'NCT', 'PMID', '(N selected)', 'Arms', 'Treatment', 'Control', 'Et', 'Nt', 'Ec', 'Nc', 'Et', 'Ec', 'Et', 'Ec', 'Et', 'Ec' ]
    sheet_data = []
    for key, value in select_extract.items():
        sheet_dict = {}
        sheet_dict['sheet_name'] = key
        sheet_dict['headers'] = csv_header
        extracted_data = [] 
        extracted_data.append(csv_second_header)
        for key_, value_ in value.items():
            # Extract data from "attrs" -> "main" -> "g0"
            if "attrs" in value_ and "main" in value_["attrs"] and "g0" in value_["attrs"]["main"]:
                data__ = value_["attrs"]["main"]["g0"]
                extracted_data.append( ['', '', key_, '', value_.get('n_arms', ''),  
                data__.get('treatment', ''), data__.get('control', ''),
                data__.get('GA_Et', ''), data__.get('GA_Nt', ''),
                data__.get('GA_Ec', ''), data__.get('GA_Nc', ''),
                data__.get('G34_Et', ''), data__.get('G34_Et', ''),
                data__.get('G3H_Et', ''), data__.get('G3H_Ec', ''), 
               data__.get('G5N_Et', ''), data__.get('G5N_Ec', '')] )

            if value_['attrs']['other']:
                data__ = value_["attrs"]["other"][0]['g0']
                extracted_data.append( ['', '', '', '', '',  
                data__.get('treatment', ''), data__.get('control', ''),
                data__.get('GA_Et', ''), data__.get('GA_Nt', ''),
                data__.get('GA_Ec', ''), data__.get('GA_Nc', ''),
                data__.get('G34_Et', ''), data__.get('G34_Et', ''),
                data__.get('G3H_Et', ''), data__.get('G3H_Ec', ''), 
                data__.get('G5N_Et', ''), data__.get('G5N_Ec', '')] )
                if len(value_["attrs"]["other"])>1:
                    data__ = value_["attrs"]["other"][1]['g0']
                    extracted_data.append( ['', '', '', '', '',  
                    data__.get('treatment', ''), data__.get('control', ''),
                    data__.get('GA_Et', ''), data__.get('GA_Nt', ''),
                    data__.get('GA_Ec', ''), data__.get('GA_Nc', ''),
                    data__.get('G34_Et', ''), data__.get('G34_Et', ''),
                    data__.get('G3H_Et', ''), data__.get('G3H_Ec', ''), 
                data__.get('G5N_Et', ''), data__.get('G5N_Ec', '')] )





        sheet_dict['data'] = extracted_data
        sheet_data.append(sheet_dict)

    
    # return response
    from openpyxl import Workbook
    wb = Workbook()

    # Remove default sheet created with workbook
    wb.remove(wb.active)

    

    # Create each sheet and populate with data
    for sheet_info in sheet_data:
        sheet_name = sheet_info['sheet_name']
        headers = sheet_info['headers']
        data = sheet_info['data']

        # Create a new sheet
        sheet = wb.create_sheet(title=sheet_name)

        # Add headers
        sheet.append(headers)

       
        for row in data:
            print(row )
            sheet.append(row)

    # Save the workbook to a temporary file
    file_path = '/tmp/output.xlsx'  # Replace with your desired path
    wb.save(file_path)

    # Serve the file for download
    return send_file(file_path, as_attachment=True)


###############################################################################
# Update the RCT sequence for a project 
###############################################################################

@bp.route('/sort_rct_seq')
@login_required
def sort_rct_seq():
    '''
    Sort the RCT origin or followup
    '''
    project_id = request.cookies.get('project_id')

    papers = dora.sort_paper_rct_seq_in_project(project_id)

    ret = {
        'success': True,
        'msg': 'sorted RCT sequences',
        'papers': [ p.as_very_simple_dict() for p in papers ]
    }
    return jsonify(ret)


###############################################################################
# Importers 
###############################################################################

@bp.route('/import_itable')
@login_required
def import_itable():
    '''
    Import or update the meta for itable

    Including:

    1. meta data for attributes
    2. data of extraction
    3. filters
    '''
    prj = request.args.get('prj')
    cq = request.args.get('cq')
    # itable = import_itable_from_xls(prj, cq)
    itable = None

    # build the return obj
    ret = {
        'success': itable is not None,
        'extract': itable.as_dict() if itable is not None else None
    }
    return jsonify(ret)


@bp.route('/import_softable_pma')
@login_required
def import_softable_pma():
    '''
    Create the meta for itable
    '''
    prj = request.args.get('prj')
    cq = request.args.get('cq')
    extracts = import_softable_pma_from_xls(prj, cq)

    ret = {
        'success': True,
        'extracts': [ ext.as_dict() for ext in extracts ]
    }

    return jsonify(ret)


###############################################################################
# Utils for the softable import
###############################################################################


def import_softable_pma_from_xls(keystr, cq_abbr, fn, group='primary'):
    '''
    Import the softable data from XLS
    '''
    if keystr == 'IO':
        return import_softable_pma_from_xls_for_IO(fn, group=group)

    raise Exception('Not implemented')


def import_softable_nma_from_xls(keystr, cq_abbr):
    '''
    Import the softable data from XLS
    '''
    raise Exception('Not implemented')


def import_softable_pma_from_xls_for_IO(fn, group='primary'):
    '''
    A special tool for importing the data file for IO project

    The data must follow this order:
    
    1. itable data / study chars 
    2. Adverse events / the category
    3. Bronchits / AE lists

    The fn could be any

    '''
    prj = 'IO'
    # ret = get_ae_pma_data(full_fn, is_getting_sms=False)
    project = dora.get_project_by_keystr(prj)
    project_id = project.project_id

    # First, create a row->pmid dictionary
    # Use the second tab
    # fn = 'ALL_DATA.xls'
    full_fn = os.path.join(
        current_app.instance_path, 
        settings.PATH_PUBDATA, 
        prj, fn
    )
    xls = pd.ExcelFile(full_fn)
    print('* load xls %s' % full_fn)
    df = xls.parse(xls.sheet_names[0], skiprows=1)
    idx2pmid = {}
    for idx, row in df.iterrows():
        idx2pmid[idx] = row['PMID']

    # Second, create tab->cate+tab dictionary
    dft = xls.parse(xls.sheet_names[1])
    ae_dict = {}
    ae_list = []
    for col in dft.columns:
        ae_cate = col
        ae_names = dft[col][~dft[col].isna()]
        ae_item = {
            'ae_cate': ae_cate,
            'ae_names': []
        }
        # build ae_name dict
        for ae_name in ae_names:
            # remove the white space
            ae_name = ae_name.strip()
            if ae_name in ae_dict:
                cate1 = ae_dict[ae_name]
                print('! duplicate %s in [%s] and [%s]' % (ae_name, cate1, ae_cate))
                continue
            ae_dict[ae_name] = ae_cate
            ae_item['ae_names'].append(ae_name)

        ae_list.append(ae_item)
            
        print('* parsed ae_cate %s with %s names' % (col, len(ae_names)))
    print('* created ae_dict %s terms' % (len(ae_dict)))

    # Third, loop on tabs
    cols = ['author', 'year', 'GA_Et', 'GA_Nt', 'GA_Ec', 'GA_Nc', 
            'G34_Et', 'G34_Ec', 'G3H_Et', 'G3H_Ec', 'G5N_Et', 'G5N_Ec', 
            'treatment', 'control']

    # each sheet is an AE/OC
    # the data tab starts from 8th
    outcomes = []
    print('* start parsing %s sheets: %s' % (
        len(xls.sheet_names[2:]), 
        xls.sheet_names[2:]
    ))
    for sheet_name in xls.sheet_names[2:]:
        print('* parsing %s' % (sheet_name))
        ae_name = sheet_name

        # create an empty meta
        ae_meta = json.loads(json.dumps(settings.OC_TYPE_TPL['pwma']['default']))
        
        # update the abbr
        ae_meta['abbr'] = util.oc_abbr()

        # update the cate
        ae_meta['category'] = ae_dict[ae_name]

        # update the group
        ae_meta['group'] = group

        # update the full name
        ae_meta['full_name'] = ae_name

        # update the input format for this project?
        ae_meta['input_format'] = 'PRIM_CAT_RAW_G5'

        # update the cate_attrs
        ae_meta['cate_attrs'] = json.loads(json.dumps(settings.INPUT_FORMAT_TPL['pwma']['PRIM_CAT_RAW_G5']))

        # get the data part
        dft = xls.parse(sheet_name, skiprows=1, usecols='A:N', names=cols)
        dft = dft[~dft.author.isna()]
        ae_data = {}
        for idx, row in dft.iterrows():
            if idx in idx2pmid:
                pmid = idx2pmid[idx]
            else:
                print('* ERROR!! Study row %s %s not defined in all studies' % (
                    idx, row[cols[0]]
                ))
                continue

            is_main = False
            if pmid not in ae_data:
                # ok, new record
                ae_data[pmid] = {
                    'is_selected': True,
                    'is_checked': True,
                    'n_arms': 2,
                    'attrs': {
                        'main': {},
                        'other': []
                    }
                }
                is_main = True
            else:
                ae_data[pmid]['n_arms'] += 1
                ae_data[pmid]['attrs']['other'].append({})

            # collect all values in the col
            for col in cols[2:]:
                # the first two cols are not used
                val = row[col]
                
                if pd.isna(val): val = None

                if is_main:
                    ae_data[pmid]['attrs']['main'][col] = val
                else:
                    ae_data[pmid]['attrs']['other'][-1][col] = val

        outcomes.append([
            ae_meta, ae_data
        ])

    # finally, save all
    oc_type = 'pwma'

    extracts = []
    for oc in outcomes:
        extract = dora.create_extract(
            project_id, 
            oc_type, 
            oc[0]['abbr'], 
            oc[0],
            oc[1]
        )
        print('* imported pwma extract %s' % (
            extract.meta['full_name']
        ))
        extracts.append(extract)

    # return!
    return extracts


@bp.route('/get_extract_piece')
@login_required
def get_extract_piece():
    '''
    Get extract piece
    '''
    project_id = request.args.get('project_id')
    extract_id = request.args.get('extract_id')
    pid = request.args.get('pid')
    # extractions_edit_ids = redis_client.get('extractions_edit_ids')
    # # extractions_edit_ids = extractions_edit_ids.append(extract_id)
    # extractions_edit_ids = json.loads(extractions_edit_ids)
    # if extractions_edit_ids and extractions_edit_ids.get(project_id):
        
    #     if not (extract_id in extractions_edit_ids[project_id]):
    #         extractions_edit_ids[project_id].append(extract_id)

    # else:
    #     extractions_edit_ids[project_id] = [extract_id]
    
    # redis_client.set('extractions_edit_ids', json.dumps(extractions_edit_ids))
    extractions_edit_ids = redis_client.get('extractions_edit_ids')
    
    # extractions_edit_ids = extractions_edit_ids.append(extract_id)
    extractions_edit_ids = json.loads(extractions_edit_ids)
    # current_user.uid
    # if extractions_edit_ids and extractions_edit_ids.get(project_id):
        
    #     if not (extract_id in extractions_edit_ids[project_id]):
    #         extractions_edit_ids[project_id].append(extract_id)

    # else:
    #     extractions_edit_ids[project_id] = [extract_id]
    extractions_edit_ids[current_user.uid] = {
        project_id : extract_id
    }
    redis_client.set('extractions_edit_ids', json.dumps(extractions_edit_ids))

    print("#############################")
    print(redis_client.get('extractions_edit_ids'))
        

    # get the     
    piece = dora.get_piece_by_project_id_and_extract_id_and_pid(
        project_id,
        extract_id,
        pid
    )

    if piece is None:
        return jsonify({
            'success': False,
            'extract_edit_ids': redis_client.get('extractions_edit_ids'),
            'current_user': current_user.uid,
            'data': {
                'piece': None
            }
        })
    
    # 2023-04-04: fix the subgroup group inconsistency
    extract = dora.get_extract(extract_id)

    # need to ensure the data completeness and consistency
    if extract.meta['group'] == 'subgroup':
        # special rule for subgroup outcome.
        # Because this group needs to create g1, g2, and even more
        # so may be issue here
        flag_updated, piece = srv_extract.validate_or_update_piece_by_outcome(
            piece, extract
        )
        if flag_updated:
            # which means this piece is updated due to missing group or other?
            dora.force_update_piece(piece)
            logging.info('updatded a subg piece due to format changes %s' % (
                piece.piece_id
            ))
    else:
        # for other group, it's ok
        pass

    return jsonify({
        'success': True,
        'extract_edit_ids': redis_client.get('extractions_edit_ids'),
        'current_user': current_user.uid,
        'data': {
            'piece': piece.as_dict()
        }
    })

@bp.route('/update_extract_edits', methods=['POST'])
@login_required
def update_extract_edits():

    project_id = request.form.get('project_id')
    extract_id = request.form.get('extract_id')
    extractions_edit_ids = redis_client.get('extractions_edit_ids')

    extractions_edit_ids = json.loads(extractions_edit_ids)
    print("Before window close")
    print(extractions_edit_ids)

    if extractions_edit_ids.get(current_user.uid):
        extractions_edit_ids.pop(current_user.uid)
    print("After window close")
    print(extractions_edit_ids)
    redis_client.set('extractions_edit_ids', json.dumps(extractions_edit_ids))
    return jsonify({
        'success': True,
    })
    
    
@bp.route('/update_extract_piece', methods=['POST'])
@login_required
def update_extract_piece():
    '''
    Get extract piece
    '''

    piece_json = json.loads(request.form.get('piece'))
    logger.info("Extraction is being added ")
    logger.info(f'User_id:{current_user.uid} with Piece Id:{piece_json.get("piece_id")} and Data: {piece_json.get("data")}')

    if 'piece_id' in piece_json:
        # which means this is an existing piece
        piece = dora.update_piece_data_by_id(
            piece_json['piece_id'],
            piece_json['data']
        )
    
    else:
        # it's a new piece!
        piece = dora.create_piece(
            piece_json['project_id'],
            piece_json['extract_id'],
            piece_json['pid'],
            piece_json['data']
        )
    
    extract_id = piece_json['extract_id']
    project_id = piece_json['project_id']
    extractions_edit_ids = redis_client.get('extractions_edit_ids')
    # extractions_edit_ids = extractions_edit_ids.append(extract_id)
    extractions_edit_ids = json.loads(extractions_edit_ids)
    if extractions_edit_ids and extractions_edit_ids.get(project_id):
        
        if extract_id in extractions_edit_ids[project_id]:
            extractions_edit_ids[project_id].remove(extract_id)

    redis_client.set('extractions_edit_ids', json.dumps(extractions_edit_ids))

    return jsonify({
        'success': True,
        'msg': 'Updated extract piece',
        'data': {
            'piece': piece.as_dict()
        }
    })


@bp.route('/get_included_papers_and_selections')
@login_required
def get_included_papers_and_selections():
    '''
    Get the included papers and the decisions of selection for outcomes
    '''
    project_id = request.args.get('project_id')
    cq_abbr = request.args.get('cq_abbr')

    if cq_abbr is None:
        cq_abbr = 'default'
    
    # get this project
    project = dora.get_project(project_id)

    # get all papers
    papers = srv_paper.get_included_papers_by_cq(
        project_id, 
        cq_abbr
    )
        
    # get all extracts
    # itable, pwma, subg, nma
    extracts = dora.get_extracts_by_project_id_and_cq(
        project_id,
        cq_abbr
    )

    # update the selection relations
    papers, extracts = dora.update_papers_outcome_selections(
        project,
        papers,
        extracts
    )

    # extend the paper meta with a new attribute
    # outcome_selections
    # and make a pid -> sequence mapping
    # pid2seq = {}
    # for i, paper in enumerate(papers):
    #     pid2seq[paper.pid] = i
    #     papers[i].meta['outcome_selections'] = []

    # check each extract
    # for extract in extracts:
    #     for pid in extract.data:
    #         if pid in pid2seq:
    #             seq = pid2seq[pid]
    #             if extract.data[pid]['is_selected']:
    #                 # this paper is selected for this outcome
    #                 papers[seq].meta['outcome_selections'].append(
    #                     extract.abbr
    #                 )
    #         else:
    #             # something wrong, this study should be there
    #             # but also maybe excluded
    #             # so, just ignore
    #             pass
    
    json_papers = [ p.as_very_simple_dict() for p in papers ]
    json_extracts = [ extr.as_simple_dict() for extr in extracts ]
    # json_extracts = [ extr.as_simple_dict() for extr in extracts ]

    ret = {
        'success': True,
        'msg': '',
        'papers': json_papers,
        'extracts': json_extracts
    }
    return jsonify(ret)

@bp.route('/check_xls_files_result_table', methods=['POST'])
@login_required
def check_xls_files_result_table():
    # get all parameters first
    project_id = request.form.get('project_id')
    cq_abbr = request.form.get('cq_abbr')

    # first, get the Result file
    file_extracts = request.files['file_extracts']
    if file_extracts.filename == '':
        return jsonify({'success': False, 'msg':'No selected file for extracts'})


    print('* upload XLS of Result Table for project %s' % project_id)
    print('* upload XLS of Result Table for cq %s' % cq_abbr)
    
    print('* file_extracts = %s' % file_extracts.filename)
    if file_extracts and util.allowed_file_format(file_extracts.filename):
        # TODO may rename the filename in the future to avoid conflict names
        fn = secure_filename(file_extracts.filename)
        # make a unique file name
        fn = '%s-%s' % (str(uuid.uuid1()), fn)
        # make full name with path
        full_fn_extracts = os.path.join(current_app.config['UPLOAD_FOLDER'], fn)
        file_extracts.save(full_fn_extracts)
    else:
        return jsonify({'success': False, 'msg': 'Not supported file format'})

    # get project
    project = dora.get_project(project_id)

    # # OK, let's save it
    print('* collecting sys stdout ...')
    stdout = sys.stdout
    sys.stdout = buffer = io.StringIO()

    try:
        project = dora.get_project_by_keystr(project.keystr)
        # load data
    

        xls = pd.ExcelFile(full_fn_extracts)

        
        first_sheet_name = xls.sheet_names[0]
        df = xls.parse(first_sheet_name, header=None, nrows=2)
        df = df.dropna(axis=1, how='all')
        dft = df.T
        df_attrs = dft.rename(columns={0: 'cate', 1: 'attr'})
        cate_attrs_relations = {}
        parent_nested_realtion = {}
        comaparator_name = ''
        varaiable_name = ''
           

        for idx, row in df_attrs.iterrows():
            if type(row['cate']) != str:
                if math.isnan(row['cate']) \
                    or math.isnan(row['attr']):
                    print('* skip nan cate or attr idx %s' % idx)
                    continue


            print("* found %s | %s" % (
                row['cate'], row['attr']
            ))

            # found cate and attr
            cate = row['cate'].strip()
            attr = row['attr'].strip()
            parent_nested_realtion[attr] = cate
            if idx == 0:
                comaparator_name = row[idx]
                continue

           
            if cate_attrs_relations.get(cate): 
                attr_subs = attr.split('|')
                if len(attr_subs) == 2:
                    # old structure, only attr + sub
                    first_level = attr_subs[0].strip()
                    second_level = attr_subs[1].strip()
                    index = 0
                    is_found = False
                    for dict_ in cate_attrs_relations[cate]:
                        index =0
                        for record in dict_['subs']:
                            
                            for i, j in record.items():
                                
                                if i == first_level:
                                    is_found = True
                                    print("@@@@@@@@@@@@@@@")
                                    print(index)
                         
                                    cate_attrs_relations[cate][0]['subs'][index][first_level][0]['sub_categories'].append(second_level)

                                index = index +1
                    if not is_found:
                        dict_ = {
                        first_level : [
                            {'sub_categories': [second_level]}
                        ]}
                        
                        cate_attrs_relations[cate][0]['subs'].append(dict_)
                                
                else:
                    first_level = attr_subs[0].strip()
                    second_level = []
                    dict_ = {
                        first_level : [
                            {'sub_categories': second_level}
                        ]}
    
                    cate_attrs_relations[cate][0]['subs'].append(dict_)
                
                
            else:
                attr_subs = attr.split('|')
                if len(attr_subs) == 2:
                    first_level = attr_subs[0].strip()
                    second_level = [attr_subs[1].strip()]
     
                else:
                    first_level = attr_subs[0].strip()
                    second_level = []
                    
                cate_attrs_relations[cate] = [{
                    'subs': [
                        {first_level : [
                            {'sub_categories': second_level}
                        ]}
                    ]
                }
                ]
        cate_attrs_relations_list = []
        for key , value in cate_attrs_relations.items():
            first_levels =[]
            for subs in value[0]['subs']:
                for _key, _value in subs.items():
                    second_levels = []
                    for sub_cate in _value[0]['sub_categories']:
                            second_levels.append({
                                "abbr": util.mk_abbr(),
                                "name": sub_cate,

                            })


                    first_levels.append({
                                "abbr": util.mk_abbr(),
                                "name": _key,
                                "sub_categories": second_levels

                            })
    
            cate_attrs_relations_list.append({
                "abbr": util.mk_abbr(),
                "name": key,
                "attrs": first_levels
            })
        abbr_list = []
        for item in cate_attrs_relations_list:
            abbr = item["abbr"]
            # abbr_list.append(abbr)
            dict_ ={}
            for attr in item["attrs"]:
                sub_abbr = attr["abbr"]
                dict_[sub_abbr] = []
                # abbr_list.append(f"{abbr} -> {sub_abbr}")
                
                for sub_attr in attr["sub_categories"]:
                    sub_sub_abbr = sub_attr["abbr"]
                    dict_[sub_abbr].append(sub_sub_abbr)
                    # abbr_list.append(f"{abbr} -> {sub_abbr} -> {sub_sub_abbr}")
                if len(dict_[sub_abbr]):
                    for r in dict_[sub_abbr]:
                        abbr_list.append(f"{abbr} -> {sub_abbr} -> {r}")
                else:
                    abbr_list.append(f"{abbr} -> {sub_abbr}")
        abbr_list_updated = [] 
        for item in abbr_list:
            parts = item.split(" -> ")
            abbr_list_updated.append(parts)

       
        xls = pd.ExcelFile(full_fn_extracts)
        first_sheet_name = xls.sheet_names[0]
        df = xls.parse(first_sheet_name, skiprows=1)

        df = df.dropna(axis=1, how='all')
        result_table_records = []
        result_table_records_seq = []
        cols = df.columns
        n_cols = len(df.columns)
        for idx in range(n_cols):
            if idx ==0:
                varaiable_name = cols[idx]
                continue
            col = cols[idx]
            result_table_records_seq.append(col)
        for _, row in df.iterrows():
            from collections import OrderedDict
            dict_ = {}
            if varaiable_name in row:
                dict_[row[varaiable_name]] = OrderedDict()

            for idx in range(n_cols):
                col = cols[idx]

                # skip the pmid column
                # since we already use this column as the key
                if col in  [varaiable_name]:
                    continue
                val = row[col]
                print(varaiable_name)
                print(dict_)
                if type(val) != str:
                    if math.isnan(val) \
                        or math.isnan(val):
                        print('* skip nan cate or attr idx %s' % idx)
                        dict_[row[varaiable_name]][col] =  'NA'
                        continue
                        
                dict_[row[varaiable_name]][col] =  val
            
            result_table_records.append(dict_)

        project_settings = project.settings 
        if project_settings.get('clinical_question_result_table'):

           
            #Existing one changes 
            project_settings['clinical_question_result_table'][cq_abbr] ={

                'cate_attrs_relations': cate_attrs_relations_list,
                'parent_nested_realtion': parent_nested_realtion,
                'result_table_records_seq': result_table_records_seq,
                'comaparator_name': comaparator_name,
                'varaiable_name': varaiable_name,
                'abbr_list_updated': abbr_list_updated,
            }
            project_settings['clinical_question_result_table'][cq_abbr]['result_table_records'] = OrderedDict()
            project_settings['clinical_question_result_table'][cq_abbr]['result_table_records'] = result_table_records
    

        else:
            project_settings['clinical_question_result_table'] = {
                cq_abbr: {
                    'cate_attrs_relations': cate_attrs_relations_list,
                    'parent_nested_realtion': parent_nested_realtion,
                    'result_table_records_seq': result_table_records_seq,
                    'comaparator_name': comaparator_name,
                    'varaiable_name': varaiable_name,
                    'abbr_list_updated': abbr_list_updated
                }
            }
            project_settings['clinical_question_result_table'][cq_abbr]['result_table_records'] = OrderedDict()
            project_settings['clinical_question_result_table'][cq_abbr]['result_table_records'] = result_table_records
        


       
        print('* Great Result Table configured Successfully !!!!!!!!!!!!')
    
        
    except Exception as ex:
        print('* Runtime Exception!')
        print('* ErrorName:', ex, ex.__traceback__)
        traceback.print_exc(file=buffer)
        print('* Please check input XLS file or contact administrator')

    sys.stdout = stdout
    sys_output = buffer.getvalue()
    print('* collected sys stdout')
    print(sys_output)

    ret = {
        'success': True,
        'msg': 'OK',
        'sys_output': sys_output,
        'cate_attrs_relations': cate_attrs_relations_list,
        'parent_nested_realtion': parent_nested_realtion,
        'result_table_records_seq': result_table_records_seq,
        'comaparator_name': comaparator_name,
        'varaiable_name': varaiable_name,
        'abbr_list_updated': abbr_list_updated,
        'result_table_records': project_settings['clinical_question_result_table'][cq_abbr]['result_table_records']


        
    }

    return jsonify(ret)

@bp.route('/check_xls_files_itable', methods=['POST'])
@login_required
def check_xls_files_itable():

    project_id = request.form.get('project_id')
    cq_abbr = request.form.get('cq_abbr')

    # first, get the itable file
    file_itable = request.files['file_itable']
    if file_itable.filename == '':
        return jsonify({'success': False, 'msg':'No selected file for itable'})
    
    print('* file_itable = %s' % file_itable.filename)
    if file_itable and util.allowed_file_format(file_itable.filename):
        # TODO may rename the filename in the future to avoid conflict names
        fn = secure_filename(file_itable.filename)
        # make a unique file name
        fn = '%s-%s' % (str(uuid.uuid1()), fn)
        # make full name with path
        full_fn_itable = os.path.join(current_app.config['UPLOAD_FOLDER'], fn)
        file_itable.save(full_fn_itable)
    else:
        return jsonify({'success': False, 'msg': 'Not supported file format'})
    

      # read the first sheet
    xls = pd.ExcelFile(full_fn_itable)
    first_sheet_name = xls.sheet_names[0]
    df = xls.parse(first_sheet_name, header=None, nrows=2)

    # 2021-06-27: weird bug, read so many NaN columns
    # df = pd.read_excel(full_fn)
    df = df.dropna(axis=1, how='all')

    # convert to other shape
    dft = df.T
    df_attrs = dft.rename(columns={0: 'cate', 1: 'attr'})
    
    itable_dict = {}
    for idx, row in df_attrs.iterrows():
        vtype = 'text'

        if type(row['cate']) != str:
            if math.isnan(row['cate']) \
                or math.isnan(row['attr']):
                print('* skip nan cate or attr idx %s' % idx)
                continue


        print("* found %s | %s" % (
            row['cate'], row['attr']
        ))

        # found cate and attr
        cate = row['cate'].strip()
        attr = row['attr'].strip()
        if itable_dict.get(cate):
            itable_dict[cate].append(attr)
        else:
            itable_dict[cate] = [attr]


    sys_output = ''
    ret = {
        'itable_dict': itable_dict,
        'success': True,
        'msg': 'OK',
        'sys_output': sys_output
    }

    return jsonify(ret)


@bp.route('/upload_xls_files_itable', methods=['POST'])
@login_required
def upload_xls_files_itable():
    '''
    Upload XLS Files for iTable
    '''
    # get all parameters first
    project_id = request.form.get('project_id')
    cq_abbr = request.form.get('cq_abbr')
    itable_flag_overwrite_decision = request.form.get('itable_flag_overwrite_decision')
    itable_flag_overwrite_piece = request.form.get('itable_flag_overwrite_piece')
    itable_flag_redirect_sys_stdout = request.form.get('itable_flag_redirect_sys_stdout')

    if cq_abbr is None:
        cq_abbr = 'default'

    if itable_flag_overwrite_decision == 'true':
        itable_flag_overwrite_decision = True
    else:
        itable_flag_overwrite_decision = False

    if itable_flag_overwrite_piece == 'true':
        itable_flag_overwrite_piece = True
    else:
        itable_flag_overwrite_piece = False

    if itable_flag_redirect_sys_stdout == 'true':
        itable_flag_redirect_sys_stdout = True
    else:
        itable_flag_redirect_sys_stdout = False

    print('* upload XLS of iTable for project %s' % project_id)
    print('* upload XLS of iTable for cq %s' % cq_abbr)
    print('* itable_flag_overwrite_decision = %s' % itable_flag_overwrite_decision)
    print('* itable_flag_overwrite_piece = %s' % itable_flag_overwrite_piece)

    # first, get the itable file
    file_itable = request.files['file_itable']
    if file_itable.filename == '':
        return jsonify({'success': False, 'msg':'No selected file for itable'})
    
    print('* file_itable = %s' % file_itable.filename)
    if file_itable and util.allowed_file_format(file_itable.filename):
        # TODO may rename the filename in the future to avoid conflict names
        fn = secure_filename(file_itable.filename)
        # make a unique file name
        fn = '%s-%s' % (str(uuid.uuid1()), fn)
        # make full name with path
        full_fn_itable = os.path.join(current_app.config['UPLOAD_FOLDER'], fn)
        file_itable.save(full_fn_itable)
    else:
        return jsonify({'success': False, 'msg': 'Not supported file format'})
    
    # second, get the filter file if it shows
    file_filter = request.files['file_filter']
    print('* file_filter = %s' % file_filter.filename)
    if file_filter.filename == '':
        # it's OK
        full_fn_filter = ''
    else:
        # TODO may rename the filename in the future to avoid conflict names
        fn = secure_filename(file_filter.filename)
        # make a unique file name
        fn = '%s-%s' % (str(uuid.uuid1()), fn)
        # make full name with path
        full_fn_filter = os.path.join(current_app.config['UPLOAD_FOLDER'], fn)
        file_filter.save(full_fn_filter)

    # get project
    project = dora.get_project(project_id)

    # OK, let's save it
    print('* collecting sys stdout ...')
    if itable_flag_redirect_sys_stdout:
        stdout = sys.stdout
        sys.stdout = buffer = io.StringIO()

    try:
        srv_extract.import_itable_from_xls(
            project.keystr,
            cq_abbr,
            full_fn_itable,
            full_fn_filter,

            # flags
            flag_overwrite_decision=itable_flag_overwrite_decision,
            flag_overwrite_piece=itable_flag_overwrite_piece
        )
    except Exception as ex:
        print('* Runtime Exception!')
        print('* ErrorName:', ex, ex.__traceback__)
        if itable_flag_redirect_sys_stdout:
            traceback.print_exc(file=buffer)
        else:
            traceback.print_exc()

        print('* Please check input XLS file or contact administrator')

    if itable_flag_redirect_sys_stdout:
        sys.stdout = stdout
        sys_output = buffer.getvalue()
        print('* collected sys stdout')
        print(sys_output)
    else:
        sys_output = ''

    ret = {
        'success': True,
        'msg': 'OK',
        'sys_output': sys_output
    }

    return jsonify(ret)


@bp.route('/upload_xls_files_pwma', methods=['POST'])
@login_required
def upload_xls_files_pwma():
    '''
    Upload XLS Files for PWMA
    '''
    # get all parameters first
    project_id = request.form.get('project_id')
    cq_abbr = request.form.get('cq_abbr')
    pwma_flag_unselect_existing_pieces = request.form.get('pwma_flag_unselect_existing_pieces')
    pwma_flag_exclude_existing_outcomes_from_public = request.form.get('pwma_flag_exclude_existing_outcomes_from_public')
    pwma_flag_redirect_sys_stdout = request.form.get('pwma_flag_redirect_sys_stdout')

    oc_type = 'pwma'
    
    if cq_abbr is None:
        cq_abbr = 'default'

    if pwma_flag_unselect_existing_pieces == 'true':
        pwma_flag_unselect_existing_pieces = True
    else:
        pwma_flag_unselect_existing_pieces = False

    if pwma_flag_exclude_existing_outcomes_from_public == 'true':
        pwma_flag_exclude_existing_outcomes_from_public = True
    else:
        pwma_flag_exclude_existing_outcomes_from_public = False

    if pwma_flag_redirect_sys_stdout == 'true':
        pwma_flag_redirect_sys_stdout = True
    else:
        pwma_flag_redirect_sys_stdout = False

    print('* upload XLS of PWMA for project %s' % project_id)
    print('* upload XLS of PWMA for cq %s' % cq_abbr)
    print('* pwma_flag_unselect_existing_pieces = %s' % pwma_flag_unselect_existing_pieces)
    print('* pwma_flag_exclude_existing_outcomes_from_public = %s' % pwma_flag_exclude_existing_outcomes_from_public)
    print('* pwma_flag_redirect_sys_stdout = %s' % pwma_flag_redirect_sys_stdout)

    # first, get the itable file
    file_extracts = request.files['file_extracts']
    if file_extracts.filename == '':
        return jsonify({'success': False, 'msg':'No selected file for extracts'})
    
    print('* file_extracts = %s' % file_extracts.filename)
    if file_extracts and util.allowed_file_format(file_extracts.filename):
        # TODO may rename the filename in the future to avoid conflict names
        fn = secure_filename(file_extracts.filename)
        # make a unique file name
        fn = '%s-%s' % (str(uuid.uuid1()), fn)
        # make full name with path
        full_fn_extracts = os.path.join(current_app.config['UPLOAD_FOLDER'], fn)
        file_extracts.save(full_fn_extracts)
    else:
        return jsonify({'success': False, 'msg': 'Not supported file format'})

    # get project
    project = dora.get_project(project_id)

    # # OK, let's save it
    print('* collecting sys stdout ...')
    if pwma_flag_redirect_sys_stdout:
        stdout = sys.stdout
        sys.stdout = buffer = io.StringIO()

    try:
        srv_extract.import_extracts_from_xls(
            project.keystr,
            cq_abbr,
            oc_type,
            full_fn_extracts,
            flag_unselect_existing_pieces=pwma_flag_unselect_existing_pieces,
            flag_exclude_existing_outcomes_from_public=pwma_flag_exclude_existing_outcomes_from_public
        )
    except Exception as ex:
        print('* Runtime Exception!')
        print('* ErrorName:', ex, ex.__traceback__)
        if pwma_flag_redirect_sys_stdout:
            traceback.print_exc(file=buffer)
        else:
            traceback.print_exc()

        print('* Please check input XLS file or contact administrator')

    if pwma_flag_redirect_sys_stdout:
        sys.stdout = stdout
        sys_output = buffer.getvalue()
        print('* collected sys stdout')
        print(sys_output)
    else:
        sys_output = ''


    ret = {
        'success': True,
        'msg': 'OK',
        'sys_output': sys_output
    }

    return jsonify(ret)

@bp.route('/check_xls_files_nma', methods=['POST'])
@login_required
def check_xls_files_nma():
    '''
    Check  XLS Files for NMA
    '''
    # get all parameters first
    project_id = request.form.get('project_id')
    cq_abbr = request.form.get('cq_abbr')
   
    oc_type = 'pwma'
    
    if cq_abbr is None:
        cq_abbr = 'default'


    # first, get the itable file
    file_extracts = request.files['file_extracts']
    if file_extracts.filename == '':
        return jsonify({'success': False, 'msg':'No selected file for extracts'})
    
    print('* file_extracts = %s' % file_extracts.filename)
    if file_extracts and util.allowed_file_format(file_extracts.filename):
        # TODO may rename the filename in the future to avoid conflict names
        fn = secure_filename(file_extracts.filename)
        # make a unique file name
        fn = '%s-%s' % (str(uuid.uuid1()), fn)
        # make full name with path
        full_fn_extracts = os.path.join(current_app.config['UPLOAD_FOLDER'], fn)
        file_extracts.save(full_fn_extracts)
    else:
        return jsonify({'success': False, 'msg': 'Not supported file format'})

    # get project
    project = dora.get_project(project_id)


    try:
        cq_plots_relations , missing_tabs, missing_pids , wrong_data, total_files, sof_files, plot_files , coe_count, missing_coe_files =  srv_extract.check_nma_files(
            project.keystr,
            cq_abbr,
            oc_type,
            full_fn_extracts,
          
        )
        pass
    except Exception as ex:
        print('* Runtime Exception!')
        print('* ErrorName:', ex, ex.__traceback__)

        print('* Please check input XLS file or contact administrator')


    sys_output = ''


    ret = {
        'success': True,
        'msg': 'OK',
        'sys_output': sys_output,
        'cq_plots_relations': cq_plots_relations,
        'missing_tabs': missing_tabs,
        'missing_pids': missing_pids,
        'wrong_data': wrong_data,
        'total_files': total_files,
        'sof_files': sof_files, 
        'plot_files': plot_files,
        'coe_count': coe_count,
        'missing_coe_files': missing_coe_files
    }

    return jsonify(ret)

@bp.route('/check_xls_files_pwma', methods=['POST'])
@login_required
def check_xls_files_pwma():
    '''
    Check  XLS Files for PWMA
    '''
    # get all parameters first
    project_id = request.form.get('project_id')
    cq_abbr = request.form.get('cq_abbr')
    pwma_flag_unselect_existing_pieces = request.form.get('pwma_flag_unselect_existing_pieces')
    pwma_flag_exclude_existing_outcomes_from_public = request.form.get('pwma_flag_exclude_existing_outcomes_from_public')
    pwma_flag_redirect_sys_stdout = request.form.get('pwma_flag_redirect_sys_stdout')

    oc_type = 'pwma'
    
    if cq_abbr is None:
        cq_abbr = 'default'

    if pwma_flag_unselect_existing_pieces == 'true':
        pwma_flag_unselect_existing_pieces = True
    else:
        pwma_flag_unselect_existing_pieces = False

    if pwma_flag_exclude_existing_outcomes_from_public == 'true':
        pwma_flag_exclude_existing_outcomes_from_public = True
    else:
        pwma_flag_exclude_existing_outcomes_from_public = False

    if pwma_flag_redirect_sys_stdout == 'true':
        pwma_flag_redirect_sys_stdout = True
    else:
        pwma_flag_redirect_sys_stdout = False

    print('* Check  XLS of PWMA for project %s' % project_id)
    print('* Check  XLS of PWMA for cq %s' % cq_abbr)
    print('* pwma_flag_unselect_existing_pieces = %s' % pwma_flag_unselect_existing_pieces)
    print('* pwma_flag_exclude_existing_outcomes_from_public = %s' % pwma_flag_exclude_existing_outcomes_from_public)
    print('* pwma_flag_redirect_sys_stdout = %s' % pwma_flag_redirect_sys_stdout)

    # first, get the itable file
    file_extracts = request.files['file_extracts']
    if file_extracts.filename == '':
        return jsonify({'success': False, 'msg':'No selected file for extracts'})
    
    print('* file_extracts = %s' % file_extracts.filename)
    if file_extracts and util.allowed_file_format(file_extracts.filename):
        # TODO may rename the filename in the future to avoid conflict names
        fn = secure_filename(file_extracts.filename)
        # make a unique file name
        fn = '%s-%s' % (str(uuid.uuid1()), fn)
        # make full name with path
        full_fn_extracts = os.path.join(current_app.config['UPLOAD_FOLDER'], fn)
        file_extracts.save(full_fn_extracts)
    else:
        return jsonify({'success': False, 'msg': 'Not supported file format'})

    # get project
    project = dora.get_project(project_id)

    # # OK, let's save it
    print('* collecting sys stdout ...')
    if pwma_flag_redirect_sys_stdout:
        stdout = sys.stdout
        sys.stdout = buffer = io.StringIO()

    try:
        cq_plots_relations , missing_tabs, missing_pids , wrong_data, total_files, sof_files, plot_files  =  srv_extract.check_pwma_files(
            project.keystr,
            cq_abbr,
            oc_type,
            full_fn_extracts,
            flag_unselect_existing_pieces=pwma_flag_unselect_existing_pieces,
            flag_exclude_existing_outcomes_from_public=pwma_flag_exclude_existing_outcomes_from_public
        )
        pass
    except Exception as ex:
        print('* Runtime Exception!')
        print('* ErrorName:', ex, ex.__traceback__)
        if pwma_flag_redirect_sys_stdout:
            traceback.print_exc(file=buffer)
        else:
            traceback.print_exc()

        print('* Please check input XLS file or contact administrator')

    if pwma_flag_redirect_sys_stdout:
        sys.stdout = stdout
        sys_output = buffer.getvalue()
        print('* collected sys stdout')
        print(sys_output)
    else:
        sys_output = ''


    ret = {
        'success': True,
        'msg': 'OK',
        'sys_output': sys_output,
        'cq_plots_relations': cq_plots_relations,
        'missing_tabs': missing_tabs,
        'missing_pids': missing_pids,
        'wrong_data': wrong_data,
        'total_files': total_files,
        'sof_files': sof_files, 
        'plot_files': plot_files
    }

    return jsonify(ret)


@bp.route('/upload_xls_result_table', methods=['POST'])
@login_required
def upload_xls_result_table():
    '''
    Upload XLS Files for Result Table
    '''
     # get all parameters first
    project_id = request.form.get('project_id')
    cq_abbr = request.form.get('cq_abbr')

    # first, get the Result file
    file_extracts = request.files['file_extracts']
    if file_extracts.filename == '':
        return jsonify({'success': False, 'msg':'No selected file for extracts'})


    print('* upload XLS of Result Table for project %s' % project_id)
    print('* upload XLS of Result Table for cq %s' % cq_abbr)
    
    print('* file_extracts = %s' % file_extracts.filename)
    if file_extracts and util.allowed_file_format(file_extracts.filename):
        # TODO may rename the filename in the future to avoid conflict names
        fn = secure_filename(file_extracts.filename)
        # make a unique file name
        fn = '%s-%s' % (str(uuid.uuid1()), fn)
        # make full name with path
        full_fn_extracts = os.path.join(current_app.config['UPLOAD_FOLDER'], fn)
        file_extracts.save(full_fn_extracts)
    else:
        return jsonify({'success': False, 'msg': 'Not supported file format'})

    # get project
    project = dora.get_project(project_id)

    # # OK, let's save it
    print('* collecting sys stdout ...')
    stdout = sys.stdout
    sys.stdout = buffer = io.StringIO()

    try:
        project = dora.get_project_by_keystr(project.keystr)
        # load data
    

        xls = pd.ExcelFile(full_fn_extracts)

        
        first_sheet_name = xls.sheet_names[0]
        df = xls.parse(first_sheet_name, header=None, nrows=2)
        df = df.dropna(axis=1, how='all')
        dft = df.T
        df_attrs = dft.rename(columns={0: 'cate', 1: 'attr'})
        cate_attrs_relations = {}
        parent_nested_realtion = {}
        comaparator_name = ''
        varaiable_name = ''
           

        for idx, row in df_attrs.iterrows():
            if type(row['cate']) != str:
                if math.isnan(row['cate']) \
                    or math.isnan(row['attr']):
                    print('* skip nan cate or attr idx %s' % idx)
                    continue


            print("* found %s | %s" % (
                row['cate'], row['attr']
            ))

            # found cate and attr
            cate = row['cate'].strip()
            attr = row['attr'].strip()
            parent_nested_realtion[attr] = cate
            if idx == 0:
                comaparator_name = row[idx]
                continue

           
            if cate_attrs_relations.get(cate): 
                attr_subs = attr.split('|')
                if len(attr_subs) == 2:
                    # old structure, only attr + sub
                    first_level = attr_subs[0].strip()
                    second_level = attr_subs[1].strip()
                    index = 0
                    is_found = False
                    for dict_ in cate_attrs_relations[cate]:
                        index =0
                        for record in dict_['subs']:
                            
                            for i, j in record.items():
                                
                                if i == first_level:
                                    is_found = True
                                    print("@@@@@@@@@@@@@@@")
                                    print(index)
                         
                                    cate_attrs_relations[cate][0]['subs'][index][first_level][0]['sub_categories'].append(second_level)

                                index = index +1
                    if not is_found:
                        dict_ = {
                        first_level : [
                            {'sub_categories': [second_level]}
                        ]}
                        
                        cate_attrs_relations[cate][0]['subs'].append(dict_)
                                
                else:
                    first_level = attr_subs[0].strip()
                    second_level = []
                    dict_ = {
                        first_level : [
                            {'sub_categories': second_level}
                        ]}
    
                    cate_attrs_relations[cate][0]['subs'].append(dict_)
                
                
            else:
                attr_subs = attr.split('|')
                if len(attr_subs) == 2:
                    first_level = attr_subs[0].strip()
                    second_level = [attr_subs[1].strip()]
     
                else:
                    first_level = attr_subs[0].strip()
                    second_level = []
                    
                cate_attrs_relations[cate] = [{
                    'subs': [
                        {first_level : [
                            {'sub_categories': second_level}
                        ]}
                    ]
                }
                ]
        cate_attrs_relations_list = []
        for key , value in cate_attrs_relations.items():
            first_levels =[]
            for subs in value[0]['subs']:
                for _key, _value in subs.items():
                    second_levels = []
                    for sub_cate in _value[0]['sub_categories']:
                            second_levels.append({
                                "abbr": util.mk_abbr(),
                                "name": sub_cate,

                            })


                    first_levels.append({
                                "abbr": util.mk_abbr(),
                                "name": _key,
                                "sub_categories": second_levels

                            })
    
            cate_attrs_relations_list.append({
                "abbr": util.mk_abbr(),
                "name": key,
                "attrs": first_levels
            })
        abbr_list = []
        for item in cate_attrs_relations_list:
            abbr = item["abbr"]
            # abbr_list.append(abbr)
            dict_ ={}
            for attr in item["attrs"]:
                sub_abbr = attr["abbr"]
                dict_[sub_abbr] = []
                # abbr_list.append(f"{abbr} -> {sub_abbr}")
                
                for sub_attr in attr["sub_categories"]:
                    sub_sub_abbr = sub_attr["abbr"]
                    dict_[sub_abbr].append(sub_sub_abbr)
                    # abbr_list.append(f"{abbr} -> {sub_abbr} -> {sub_sub_abbr}")
                if len(dict_[sub_abbr]):
                    for r in dict_[sub_abbr]:
                        abbr_list.append(f"{abbr} -> {sub_abbr} -> {r}")
                else:
                    abbr_list.append(f"{abbr} -> {sub_abbr}")
        abbr_list_updated = [] 
        for item in abbr_list:
            parts = item.split(" -> ")
            abbr_list_updated.append(parts)

       
        xls = pd.ExcelFile(full_fn_extracts)
        first_sheet_name = xls.sheet_names[0]
        df = xls.parse(first_sheet_name, skiprows=1)

        df = df.dropna(axis=1, how='all')
        result_table_records = []
        result_table_records_seq = []
        cols = df.columns
        n_cols = len(df.columns)
        for idx in range(n_cols):
            if idx ==0:
                varaiable_name = cols[idx]
                continue
            col = cols[idx]
            result_table_records_seq.append(col)
        for _, row in df.iterrows():
            from collections import OrderedDict
            dict_ = {}
            if varaiable_name in row:
                dict_[row[varaiable_name]] = OrderedDict()

            for idx in range(n_cols):
                col = cols[idx]

                # skip the pmid column
                # since we already use this column as the key
                if col in  [varaiable_name]:
                    continue
                val = row[col]
                print(varaiable_name)
                print(dict_)
                if type(val) != str:
                    if math.isnan(val) \
                        or math.isnan(val):
                        print('* skip nan cate or attr idx %s' % idx)
                        dict_[row[varaiable_name]][col] =  'NA'
                        continue
                        
                dict_[row[varaiable_name]][col] =  val
            
            result_table_records.append(dict_)

        project_settings = project.settings 
        if project_settings.get('clinical_question_result_table'):

           
            #Existing one changes 
            project_settings['clinical_question_result_table'][cq_abbr] ={

                'cate_attrs_relations': cate_attrs_relations_list,
                'parent_nested_realtion': parent_nested_realtion,
                'result_table_records_seq': result_table_records_seq,
                'comaparator_name': comaparator_name,
                'varaiable_name': varaiable_name,
                'abbr_list_updated': abbr_list_updated,
            }
            project_settings['clinical_question_result_table'][cq_abbr]['result_table_records'] = OrderedDict()
            project_settings['clinical_question_result_table'][cq_abbr]['result_table_records'] = result_table_records
    

        else:
            project_settings['clinical_question_result_table'] = {
                cq_abbr: {
                    'cate_attrs_relations': cate_attrs_relations_list,
                    'parent_nested_realtion': parent_nested_realtion,
                    'result_table_records_seq': result_table_records_seq,
                    'comaparator_name': comaparator_name,
                    'varaiable_name': varaiable_name,
                    'abbr_list_updated': abbr_list_updated
                }
            }
            project_settings['clinical_question_result_table'][cq_abbr]['result_table_records'] = OrderedDict()
            project_settings['clinical_question_result_table'][cq_abbr]['result_table_records'] = result_table_records
        


        is_success, project = dora.set_project_settings(
            project_id, project_settings
        )
        print('* Great Result Table configured Successfully !!!!!!!!!!!!')
    
        
    except Exception as ex:
        print('* Runtime Exception!')
        print('* ErrorName:', ex, ex.__traceback__)
        traceback.print_exc(file=buffer)
        print('* Please check input XLS file or contact administrator')

    sys.stdout = stdout
    sys_output = buffer.getvalue()
    print('* collected sys stdout')
    print(sys_output)

    ret = {
        'success': True,
        'msg': 'OK',
        'sys_output': sys_output
    }

    return jsonify(ret)


@bp.route('/upload_xls_files_nma', methods=['POST'])
@login_required
def upload_xls_files_nma():
    '''
    Upload XLS Files for NMA
    '''
    # get all parameters first
    project_id = request.form.get('project_id')
    cq_abbr = request.form.get('cq_abbr')
    nma_flag_unselect_existing_pieces = request.form.get('nma_flag_unselect_existing_pieces')
    nma_flag_exclude_existing_outcomes_from_public = request.form.get('nma_flag_exclude_existing_outcomes_from_public')
    nma_flag_has_coe = request.form.get('nma_flag_has_coe')
    oc_type = 'nma'
    
    if cq_abbr is None:
        cq_abbr = 'default'

    if nma_flag_unselect_existing_pieces == 'true':
        nma_flag_unselect_existing_pieces = True
    else:
        nma_flag_unselect_existing_pieces = False

    if nma_flag_exclude_existing_outcomes_from_public == 'true':
        nma_flag_exclude_existing_outcomes_from_public = True
    else:
        nma_flag_exclude_existing_outcomes_from_public = False

    if nma_flag_has_coe == 'true':
        nma_flag_has_coe = True
    else:
        nma_flag_has_coe = False

    print('* upload XLS of NMA for project %s' % project_id)
    print('* upload XLS of NMA for cq %s' % cq_abbr)
    print('* nma_flag_unselect_existing_pieces = %s' % nma_flag_unselect_existing_pieces)
    print('* nma_flag_exclude_existing_outcomes_from_public = %s' % nma_flag_exclude_existing_outcomes_from_public)
    print('* nma_flag_has_coe = %s' % nma_flag_has_coe)

    # first, get the itable file
    file_extracts = request.files['file_extracts']
    if file_extracts.filename == '':
        return jsonify({'success': False, 'msg':'No selected file for extracts'})
    
    print('* file_extracts = %s' % file_extracts.filename)
    if file_extracts and util.allowed_file_format(file_extracts.filename):
        # TODO may rename the filename in the future to avoid conflict names
        fn = secure_filename(file_extracts.filename)
        # make a unique file name
        fn = '%s-%s' % (str(uuid.uuid1()), fn)
        # make full name with path
        full_fn_extracts = os.path.join(current_app.config['UPLOAD_FOLDER'], fn)
        file_extracts.save(full_fn_extracts)
    else:
        return jsonify({'success': False, 'msg': 'Not supported file format'})

    # get project
    project = dora.get_project(project_id)

    # # OK, let's save it
    print('* collecting sys stdout ...')
    stdout = sys.stdout
    sys.stdout = buffer = io.StringIO()

    try:
        srv_extract.import_extracts_from_xls(
            project.keystr,
            cq_abbr,
            oc_type,
            full_fn_extracts,
            flag_unselect_existing_pieces=nma_flag_unselect_existing_pieces,
            flag_exclude_existing_outcomes_from_public=nma_flag_exclude_existing_outcomes_from_public,
            flag_has_nma_coe=nma_flag_has_coe,
        )
    except Exception as ex:
        print('* Runtime Exception!')
        print('* ErrorName:', ex, ex.__traceback__)
        traceback.print_exc(file=buffer)
        print('* Please check input XLS file or contact administrator')

    sys.stdout = stdout
    sys_output = buffer.getvalue()
    print('* collected sys stdout')
    print(sys_output)
    ret = {
        'success': True,
        'msg': 'OK',
        'sys_output': sys_output
    }

    return jsonify(ret)